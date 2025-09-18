from typing import Optional, List
from unittest.mock import Base
from langchain.callbacks.manager import (
    CallbackManagerForToolRun,
)
import os
from pydantic import BaseModel, Field

from langchain_core.tools.base import BaseTool
from regex import F

from schemas import *

import pandas as pd

import logging

import sys
if sys.version_info[0] < 3: 
    from StringIO import StringIO
else:
    from io import StringIO

import openpyxl 
from openpyxl.worksheet.worksheet import Worksheet

import excel

logger = logging.getLogger(__name__)

class UnzipFileTool(BaseTool):
    name: str = "Unzip"
    description: str = (
        """
            Descompacta um determinado arquivo a partir de uma determinada pasta. 
            Retorna uma lista contendo os caminhos dos arquivos descompactados.
        """
    )
    args_schema: type[BaseModel] = UnzipFileInput
    return_direct: bool = False

    def _run(
        self,
        nome_arquivo: str,
        diretorio: str,
        run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> List[str]:
        """Utiliza a ferramenta de forma síncrona."""
        import zipfile
        from pathlib import Path

        logging.info(f"Descomprimindo o arquivo {nome_arquivo} dentro da pasta {diretorio}...")

        full_path = os.path.join(diretorio, nome_arquivo)

        with zipfile.ZipFile(full_path, "r") as zip_ref:
            zip_ref.extractall(diretorio)

        directory_path = Path(diretorio)

        paths_arquivos_descompactados = [os.path.join(diretorio,entry.name) for entry in directory_path.iterdir() if entry.is_file()]

        logging.info(f"Os arquivos descompactados são: {paths_arquivos_descompactados}")

        return paths_arquivos_descompactados
    
class ReunirDadosTool(BaseTool):

    name: str = "ReunirDados"
    description: str = "Reune ou concatena os dados de uma ou mais planilhas. Retorna o caminho da planilha em Excel cujos dados foram mesclados."
    return_direct: bool = False
    args_schema: type[BaseModel] = ReunirDadosInput

    def _run(self,
        paths: list[str] | str,
        run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> str:
        
        if isinstance(paths,list):
            paths_planilhas_excel = paths
        else:
            paths_planilhas_excel = [ _.strip() for _ in paths.split(",")]
        
        logger.info(f"Reunindo os dados oriundos das planilhas localizadas em {paths_planilhas_excel} (total de {len(paths_planilhas_excel)} arquivos) ...")

        return excel.mesclar(paths_planilhas_excel,PlanilhaTemporaria().obter_caminho_arquivo_temporario("merged.xlsx"))
    
"""
class CalcularQuantidadeDiasUteisTool(BaseTool):
    name: str = "CalcularQuantidadeDiasUteis"
    description: str = "Calcula a quantidade de dias úteis por colaborador considerando como base os dias úteis por sindicato."
    return_direct: bool = False
    args_schema: type[BaseModel] = CalcularQuantidadeDiasUteisInput

    def _run(self,
        paths: list[str] | str,
        dias_uteis_por_sindicato: dict[str,str] | str,
        run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> int:
        
        if isinstance(paths,list):
            paths_planilhas_excel = paths
        else:
            paths_planilhas_excel = [ _.strip() for _ in paths.split(",")]

        logger.info(f"Calculando os dias úteis a partir das planilhas {paths} e dias úteis por sindicato {dias_uteis_por_sindicato} ...")

        return 0
"""
    
"""
class ExtrairDadosColunasTool(BaseTool):
    name: str = "ExtrairDadosColunas"
    description: str = "Extrai ou obtém os dados das colunas de uma determinada planilha. Retorna o caminho da planilha em Excel cujos dados das colunas foram extraídos."
    return_direct: bool = False
    args_schema: type[BaseModel] = ExtrairDadosColunasInput

    def _run(self,
             path: str,
             nomes_colunas: list[str] | str,
            run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> str:
        
        if isinstance(nomes_colunas,list):
            colunas_filtradas = nomes_colunas
        else:
            colunas_filtradas = [ _.strip() for _ in nomes_colunas.split(",")]

        logger.info(f"Extraindo os dados das colunas {colunas_filtradas} da planilha {path}...")

        df = pd.read_excel(path,index_col=0)

        df = df[colunas_filtradas]

        df.to_excel(path)

        return path
"""
"""
class ObterDadosTool(BaseTool):
    name: str = "ObterDados"
    description: str = "Obtem os dados de uma planilha determinada. Retorna o numero de linhas que esta planilha possui."
    return_direct: bool = False
    args_schema: type[BaseModel] = ObterDadosInput

    def _run(self,
        path: str,
        run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> int:
        
        logger.info(f"Obtendo os dados da planilha {path} ...")

        dados = pd.read_excel(path)

        logger.info(f"Total de linhas: {dados.shape[0]}")

        return dados.shape[0]
"""

class EscreverDadosNaPlanilhaTool(BaseTool):
    name: str = "EscreverDadosNaPlanilhaTool"
    description: str = "Escreve ou copia os dados em uma planilha destino, completando a mesma."
    return_direct: bool = True
    args_schema: type[BaseModel] = EscreverDadosNaPlanilhaInput

    def _run(
        self,
        path_origem: str,
        path_destino: str,
        competencia: str,
#        nome_aba_planilha_destino: Optional[str] = None,
        run_manager: Optional[CallbackManagerForToolRun] = None,
    ) -> str:
        

        logger.info(f"Exportando os dados oriundos da planilha {path_origem} para a planilha final em {path_destino} e competencia {competencia}...")

        wb_destino = openpyxl.load_workbook(path_destino)

        ws_destino = wb_destino.active

        wb_origem = openpyxl.load_workbook(path_origem)
        ws_origem = wb_origem.active

        #Agora, para cada coluna header de origem, vamos procurar a posicao correspondente na coluna destino:
        header_destino = [cell.value for cell in ws_destino[2] if ws_destino[2]!='Unnamed']
        header_origem = [cell.value for cell in ws_origem[1] if ws_origem[1]!='Unnamed']

        logger.debug(f"Header origem = {header_origem} e header destino = {header_destino}")

        i = self.__index_first_empty_row(ws_destino)

        num_primeira_linha_nao_preenchida = i

        for row in ws_origem.iter_rows(min_row=2):
            for j_origem in range(len(row)):
                if header_origem[j_origem]:
                    j_destino = self.__index_matching_col(header_origem[j_origem],header_destino)
                    if j_destino != None:
                        cell_origem = row[j_origem]
                        logger.debug(f"Escrevendo na planilha destino o valor {cell_origem.value} na linha {i} e coluna {j_destino+1}...")
                        ws_destino.cell(row=i,column=j_destino+1,value=cell_origem.value)
            i += 1

        # Preenche a compentência na coluna Competencia:
        i = num_primeira_linha_nao_preenchida
        j = self.__index_matching_col("Competencia",header_destino)
        for row in ws_destino.iter_rows(min_row=num_primeira_linha_nao_preenchida):
            if row[0].value:
                ws_destino.cell(row=i,column=j+1,value=competencia.replace(".","/"))
            i +=1

        assert j!=None

        # Vamos definir a formula total:
        ws_destino['G1'].value = f"=SUM($G{num_primeira_linha_nao_preenchida-1}:$G{ws_destino.max_row})"

        excel.autofit(ws_destino)

        wb_destino.save(path_destino)

        logger.info(f"Planilha {path_destino} preenchida.")

        return path_destino
    

    def __index_matching_col(self, word:str, header:list[str]) -> Optional[int]:
        from fuzzywuzzy import fuzz
        logger.debug(f"Buscando o índice no header {header} com outra palavra que se case com {word}...")
        for j in range (len(header)):
            header_word = header[j]
            score = fuzz.partial_ratio(header_word.lower(),word.lower())
            logger.debug(f"Score de similaridade entre as palavras {header_word} e {word}: {score}")
            if (score >= 80):
                logger.debug(f"As palavras {header_word} e {word} casaram! O índice correspondente na lista é {j}!")
                return j
        
        logger.debug(f"Infelizmente não localizei similaridade da palavra {word} com alguma da lista {header}!")
        return None

            

    def __index_first_empty_row(self,ws: type[Worksheet]) -> int:
        for row_index in range(1, ws.max_row + 2):  # Iterate up to one row beyond max_row
            is_empty = True
            for col_index in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_index, column=col_index).value
                if cell_value is not None:
                    is_empty = False
                    break  # Found a non-empty cell in this row, so it's not empty
            if is_empty:
                return row_index

import tempfile
import shutil
import atexit
class PlanilhaTemporaria:

    def __init__(self):
        self.__temp_dir = tempfile.mkdtemp()
        atexit.register(self.__cleanup_function, "Closing files")

    def obter_caminho_arquivo_temporario(self,filename:str) -> str:
#        return  os.path.join('data/',filename)
        return  os.path.join(self.__temp_dir,filename)

    def exportar_dados_planilha_temporaria(self,df: type[pd.DataFrame]) -> str:
        excel_destino =  os.path.join(self.__temp_dir,"output.xlsx")
        df.to_excel(excel_destino,index=True,index_label='Matricula')
        logger.info(f"Os dados foram escritos com sucesso em {excel_destino}")
        return excel_destino    

    def __cleanup_function(self,message):
        logger.info(message)
        try:
           shutil.rmtree(self.__temp_dir)
        except IOError:
            sys.stderr.write('Failed to clean up temp dir {}'.format(self.__temp_dir))
