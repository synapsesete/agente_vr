from typing import Optional, Tuple
from importlib.resources import path
import logging
import openpyxl.worksheet
import pandas as pd
import openpyxl 
from openpyxl.worksheet.worksheet import Worksheet
from fuzzywuzzy import fuzz

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)

logger = logging.getLogger(__name__)


def autofit(worksheet:type[Worksheet],scale=1.0) -> None:
    for col in worksheet.columns:
     max_length = 0
     column = col[0].column_letter # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length ) * scale
     worksheet.column_dimensions[column].width = adjusted_width

def mesclar(planilhas_excel: list[str],dest_filename:str = "output.xlsx") -> Optional[str]:
    """
     A partir de uma lista de caminhos de arquivos em Excel, mescla todas elas retornando o caminho do arquivo
     gerado.
    """
    logger.info(f"Mesclando as planilhas {planilhas_excel} através do Pandas (total de planilhas: {len(planilhas_excel)})...")

    if len(planilhas_excel)==0:
        return None
    elif len(planilhas_excel)==1:
        return planilhas_excel[0]
    else:
        planilha_merged = planilhas_excel[0]
        for i in range(1,len(planilhas_excel)):
            planilha1 = planilha_merged
            planilha2 = planilhas_excel[i]
            tuplas = buscar_indices_merging_por_similaridade(planilha1,planilha2)
            if tuplas:
                tupla1 = tuplas[0]
                df1 = pd.read_excel(planilha1,skiprows=tupla1[0]-1,index_col=tupla1[1]-1)
                tupla2 = tuplas[1]
                df2 = pd.read_excel(planilha2,skiprows=tupla2[0]-1,index_col=tupla2[1]-1)
                logger.info("Os nomes dos índices dos dataframes criados na primeira planilha %s é %s e na segunda planilha %s é %s",planilha1,df1.index.name,planilha2,df2.index.name)
                if df1.index.name == df2.index.name:
                    merged = df1.combine_first(df2)
                else:
                    merged = df1.join(df2)
            else:
                df1 = pd.read_excel(planilha1)
                df2 = pd.read_excel(planilha2)
                merged = pd.concat([df1,df2],axis=0)

            planilha_merged = dest_filename
            merged.to_excel(planilha_merged)
        
        return dest_filename


def buscar_indices_merging_por_similaridade(arq_excel1: str, arq_excel2: str) -> Optional[Tuple[Tuple[int,int],Tuple[int,int]]]:
    """
        Busca duas tuplas de indices (i,j) para as duas planilhas passadas referentes à similaridade de nome de coluna.
    """
    wb_obj1 = openpyxl.load_workbook(arq_excel1)    
    wb_obj2 = openpyxl.load_workbook(arq_excel2)    

    sheet1 = wb_obj1.active
    sheet2 = wb_obj2.active

    for i1 in range(1,sheet1.max_row+1):
        for j1 in range(1,sheet1.max_column+1):
            cell1 = sheet1.cell(row=i1,column=j1)
            for i2 in range(1,sheet2.max_row+1):
                for j2 in range(1,sheet2.max_column+1):
                    cell2 = sheet2.cell(row=i2,column=j2)
                    if cell1.data_type=='s' and cell2.data_type=='s':
                        score = fuzz.ratio(cell1.value.lower(),cell2.value.lower())
                        if score>=80:
                            logger.info("As palavras %s e %s possuem o score %d (arquivos %s e %s)!",cell1.value.lower(),cell2.value.lower(),score,arq_excel1,arq_excel2)
                            logger.info("Achei por similaridade as coordenadas (%d,%d) e (%d,%d)!",i1,j1,i2,j2)
                            return ((i1,j1),(i2,j2))


