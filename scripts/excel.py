import logging
from typing import Optional, Tuple

import openpyxl
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def remover_registros_planilha_por_valores_especificos_coluna(planilha_origem: str,token: str,valores:list[str],indice_planilha_origem=1) -> str:

    wb_origem = openpyxl.load_workbook(planilha_origem)
    ws_origem = wb_origem.active
    __remover_registros_worksheets_por_valores_especificos_coluna(ws_origem,token,valores,indice_planilha_origem)

    wb_origem.save(planilha_origem)

    return planilha_origem
    

def __remover_registros_worksheets_por_valores_especificos_coluna(ws_origem: type[Worksheet], token: str,valores:list[str],indice_planilha_origem=1) -> None:

    first_row_origem = ws_origem[indice_planilha_origem]

    indices = buscar_todos_indices_row_por_similaridade(token,first_row_origem)

    if indices:
        logger.debug("Todos os indices: %s",indices)
        for i in indices:
            nome_coluna = first_row_origem[i].value
            logger.debug("O nome da coluna que encontrei se chama %s",nome_coluna)
            for row_index in range(indice_planilha_origem+1, ws_origem.max_row+1):
                valor_celula = ws_origem.cell(row=row_index,column=i+1).value
                logger.debug("valor da celula [%d,%d] = %s",row_index,i+1,valor_celula)
                if valor_celula!=None:
                    for valor in valores:
                        score = fuzz.ratio(valor.strip().lower(), str(valor_celula).strip().lower())
                        if (score>50):
                            logger.debug("O valor %s casa com o valor %s!",valor_celula,valor)
                            ws_origem.delete_rows(row_index,1)
                            __remover_registros_worksheets_por_valores_especificos_coluna(ws_origem,token,valores,indice_planilha_origem)
                            return 
        



def preencher_planilha(
    planilha_origem: str,
    planilha_destino: str,
    indice_planilha_origem=1,
    indice_planilha_destino=1,
    max_col_planilha_destino=None,
) -> str:

    logger.info(
        f"Preenchendo os dados oriundos da planilha {planilha_origem} na planilha final em {planilha_destino} ..."
    )

    wb_destino = openpyxl.load_workbook(planilha_destino)
    ws_destino = wb_destino.active

    wb_origem = openpyxl.load_workbook(planilha_origem)
    ws_origem = wb_origem.active

    i = index_first_empty_row(ws_destino)
    num_primeira_linha_nao_preenchida = i
    logger.debug(
        "Primeira linha nao preenchida na planilha destino é %d",
        num_primeira_linha_nao_preenchida,
    )

    first_row_origem = ws_origem[indice_planilha_origem]

    if max_col_planilha_destino:
        first_row_destino = ws_destino[indice_planilha_destino][
            0:max_col_planilha_destino
        ]
    else:
        first_row_destino = ws_destino[indice_planilha_destino]

    for j_origem in range(len(first_row_origem)):
        cell_origem = first_row_origem[j_origem]
        j_destino = buscar_indice_row_por_similaridade(
            cell_origem.value, first_row_destino
        )
        if j_destino != None:
            logger.debug(
                "Copiando os dados da coluna de origem de nome %s para a coluna destino de nome %s",
                cell_origem.value,
                first_row_destino[j_destino].value,
            )
            __copiar_coluna(
                ws_origem,
                ws_destino,
                j_origem + 1,
                j_destino + 1,
                2,
                num_primeira_linha_nao_preenchida,
            )

    wb_destino.save(planilha_destino)

def buscar_todos_indices_row_por_similaridade(word: str, row: type[tuple]) -> Optional[tuple]:

    indices = []

    for j in range(len(row)):
        cell = row[j]
        logger.debug("Valor da celula: %s", cell.value)
        if cell.data_type == "s":
            word_cell = cell.value
            logger.debug(
                "Verificando se a palavra %s da linha destino casa com a palavra %s...",
                word_cell,
                word,
            )
            score = fuzz.partial_ratio(
                word.strip().lower(), str(word_cell).strip().lower()
            )
            if score > 90:
                score_full = fuzz.ratio(
                    word.strip().lower(), str(word_cell).strip().lower()
                )
                logger.debug(
                    "As palavras %s e %s possuem o score %d e score-full %d. O indice da coluna é %d",
                    word,
                    word_cell,
                    score,
                    score_full,
                    j,
                )
                indices.append(j)

    return tuple(indices)


def buscar_indice_row_por_similaridade(word: str, row: type[tuple]) -> Optional[int]:

    for j in range(len(row)):
        cell = row[j]
        logger.debug("Valor da celula: %s", cell.value)
        if cell.data_type == "s":
            word_cell = cell.value
            logger.debug(
                "Verificando se a palavra %s da linha destino casa com a palavra %s...",
                word_cell,
                word,
            )
            score = fuzz.partial_ratio(
                word.strip().lower(), str(word_cell).strip().lower()
            )
            if score > 90:
                score_full = fuzz.ratio(
                    word.strip().lower(), str(word_cell).strip().lower()
                )
                logger.debug(
                    "As palavras %s e %s possuem o score %d e score-full %d. O indice da coluna é %d",
                    word,
                    word_cell,
                    score,
                    score_full,
                    j,
                )
                return j


def __copiar_coluna(
    ws_source: type[Worksheet],
    ws_destino: type[Worksheet],
    j_source: int,
    j_destino: int,
    skiprows_source: int = 1,
    skip_rows_destino: int = 1,
) -> None:

    row_init_origem = skiprows_source
    row_fim_origem = ws_source.max_row + 1

    row_init_destino = skip_rows_destino

    logger.debug(
        "Copiando a coluna %d do ws origem para a coluna %d do ws destino a partir da linha %d:%d da planilha de origem começando na linha %d na planilha destino.",
        j_source,
        j_destino,
        row_init_origem,
        row_fim_origem,
        row_init_destino,
    )

    # Iterate through rows and copy cell values
    i = 0
    for row_index in range(row_init_origem, row_fim_origem):
        source_cell_value = ws_source.cell(row=row_index, column=j_source).value
        ws_destino.cell(row=row_init_destino + i, column=j_destino).value = (
            source_cell_value
        )
        i += 1
    #   logger.info("Celula de valor %s na origem copiada para o destino!",source_cell_value)


def index_first_empty_row(ws: type[Worksheet]) -> int:
    for row_index in range(1, ws.max_row + 2):  # Iterate up to one row beyond max_row
        is_empty = True
        for col_index in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_index, column=col_index).value
            if cell_value is not None:
                is_empty = False
                break  # Found a non-empty cell in this row, so it's not empty
        if is_empty:
            return row_index


def autofit(worksheet: type[Worksheet], scale=1.0) -> None:
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length) * scale
        worksheet.column_dimensions[column].width = adjusted_width


def mesclar(
    planilhas_excel: list[str], dest_filename: str = "output.xlsx"
) -> Optional[str]:
    """
    A partir de uma lista de caminhos de arquivos em Excel, mescla todas elas retornando o caminho do arquivo
    gerado.
    """
    logger.info(
        f"Mesclando as planilhas {planilhas_excel} através do Pandas (total de planilhas: {len(planilhas_excel)})..."
    )

    if len(planilhas_excel) == 0:
        return None
    elif len(planilhas_excel) == 1:
        return planilhas_excel[0]
    else:
        planilha_merged = planilhas_excel[0]
        for i in range(1, len(planilhas_excel)):
            planilha1 = planilha_merged
            planilha2 = planilhas_excel[i]
            tuplas = buscar_indices_merging_por_similaridade(planilha1, planilha2)
            if tuplas:
                tupla1 = tuplas[0]
                df1 = pd.read_excel(
                    planilha1, skiprows=tupla1[0] - 1, index_col=tupla1[1] - 1
                )
                tupla2 = tuplas[1]
                df2 = pd.read_excel(
                    planilha2, skiprows=tupla2[0] - 1, index_col=tupla2[1] - 1
                )
                logger.debug(
                    "Os nomes dos índices dos dataframes criados na primeira planilha %s é %s e na segunda planilha %s é %s",
                    planilha1,
                    df1.index.name,
                    planilha2,
                    df2.index.name,
                )
                if df1.index.name == df2.index.name:
                    merged = df1.combine_first(df2)
                else:
                    merged = df1.join(df2)
            else:
                df1 = pd.read_excel(planilha1)
                df2 = pd.read_excel(planilha2)
                merged = pd.concat([df1, df2], axis=0)
            planilha_merged = dest_filename

            merged.to_excel(planilha_merged)

        return dest_filename


def buscar_indices_merging_por_similaridade(
    arq_excel1: str, arq_excel2: str
) -> Optional[Tuple[Tuple[int, int], Tuple[int, int]]]:
    """
    Busca duas tuplas de indices (i,j) para as duas planilhas passadas referentes à similaridade de nome de coluna.
    """
    wb_obj1 = openpyxl.load_workbook(arq_excel1)
    wb_obj2 = openpyxl.load_workbook(arq_excel2)

    sheet1 = wb_obj1.active
    sheet2 = wb_obj2.active

    for i1 in range(1, sheet1.max_row + 1):
        for j1 in range(1, sheet1.max_column + 1):
            cell1 = sheet1.cell(row=i1, column=j1)
            for i2 in range(1, sheet2.max_row + 1):
                for j2 in range(1, sheet2.max_column + 1):
                    cell2 = sheet2.cell(row=i2, column=j2)
                    if cell1.data_type == "s" and cell2.data_type == "s":
                        word1 = str(cell1.value).strip().lower()
                        word2 = str(cell2.value).strip().lower()
                        score = fuzz.ratio(word1, word2)
                        if score > 80:
                            logger.debug(
                                "As palavras %s e %s possuem o score %d (arquivos %s e %s)!",
                                word1,
                                word2,
                                score,
                                arq_excel1,
                                arq_excel2,
                            )
                            logger.debug(
                                "Achei por similaridade as coordenadas (%d,%d) e (%d,%d)!",
                                i1,
                                j1,
                                i2,
                                j2,
                            )
                            return ((i1, j1), (i2, j2))


if __name__ == "__main__":
    remover_registros_planilha_por_valores_especificos_coluna("data/merged.xlsx",'Cargo',['desenvolvedores','estagiários','diretores'])
